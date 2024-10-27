import pandas as pd
import numpy as np
import xlwt
import openpyxl
from xlwt import Workbook

path = "C:\\Users\\1thom\\OneDrive\\Bureau\\Ultimate_Budget_Empty.xlsx"
path2 = "C:\\Users\\1thom\\Downloads\\export-operations-27-10-2024_08-51-13.csv"

df = pd.read_csv(path2, sep=';')
df_copy = df.drop(columns=['dateVal', 'categoryParent', 'comment', 'accountNum', 'accountLabel', 'accountbalance'])

df_copy['dateOp'] = pd.to_datetime(df_copy['dateOp'], errors='coerce')
# Remove commas and convert 'amount' to float
df_copy['amount'] = df_copy['amount'].str.replace(' ', '').str.replace(',', '.').astype(float)

"""# Add a 'year' column
df_copy['year'] = df_copy['dateOp'].dt.year
# Add 'month' column
df_copy['month'] = df_copy['dateOp'].dt.month

# Group by 'year' and 'month' and get the average sales
monthly_ex = df_copy.groupby(['year', 'month'])['amount'].mean()

print(monthly_ex)"""

# Sort by date
df_copy = df_copy.sort_values(by="dateOp")


wb = openpyxl.load_workbook(path)

sheet = wb["Tracking Budget"]

# Note: The first row or
# column integer is 1, not 0.

# Cell object is created by using
# sheet object's cell() method.
cell_obj = sheet.cell(row = 12, column = 4)

# Print value of cell object
# using the value attribute
print(cell_obj.value)

# print total number of column
max_col = sheet.max_column
print(max_col)
# Loop will print all columns name
"""for i in range(1, max_col + 1):
    cell_obj = sheet.cell(row=1, column=i)
    print(cell_obj.value)"""

# B2 means column = 2 & row = 2.
#c4 = sheet['K10']
#c4.value = "RAI"


print(df_copy.info())
print(df_copy.head())
#wb.save(path)

















