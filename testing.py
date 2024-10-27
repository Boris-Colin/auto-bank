import pandas as pd
import numpy as np
import xlwt
import openpyxl
from xlwt import Workbook


path = "C:\\Users\\1thom\\OneDrive\\Bureau\\Ultimate_Budget_Empty.xlsx"

wb = openpyxl.load_workbook(path)
sheet = wb["Tracking Budget"]


# Note: The first row or
# column integer is 1, not 0.

# Cell object is created by using
# sheet object's cell() method.
cell_obj = sheet.cell(row = 12, column = 4)

# Print value of cell object
# using the value attribute
print(cell_obj.value)

# Define the starting row of the table
table_start_row = 11  # Adjust this to where your table actually begins

# Find the first empty row within the table
first_empty_row = table_start_row
while sheet.cell(row=first_empty_row, column=2).value is not None:
    first_empty_row += 1

print(first_empty_row)

# B2 means column = 2 & row = 2.
#c4 = sheet['K10']
#c4.value = "RAI"

#wb.save(path)
